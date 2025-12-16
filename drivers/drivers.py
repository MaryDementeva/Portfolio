
pip install schedule # библиотека работы с выполнением действий по расписанию

!pip install PyPDF2  #используемая библиотека распознавания Pdf файлов

pip install Pillow # модуль работы с изображениями

import pandas as pd #
import sys
import os #работа с системными характеристиками файлов (имя, дата создания)
import glob #
from datetime import datetime
from datetime import date
import smtplib # библиотека для отправки сообщение с почты (логин/пароль/доступ к почте)
from email.message import EmailMessage #библиотека для формирования самого почтового сообщения (тема, вложение, сопутствующее сообщение)
import schedule # библиотека для выполнения действий по расписанию
import time
from PyPDF2 import PdfReader # метод чтения pdf файлов (текст + изображение)
import re # разбор текста по шаблонам
import matplotlib.pyplot as plt
from PIL import Image
import cv2 #для считывания файлов как картинок

import numpy as np

"""**Библиотека и ее модули для записи в эксель**"""

import openpyxl #сама библиотека
from openpyxl import Workbook # модуль для создания нового эксель файла
from openpyxl.drawing.image import Image #модуль для записи картинок в эксель
from openpyxl.styles import Font #модуль для изменения параметров ячейки
from openpyxl.styles import Alignment #модуль для переноса текста в ячейке
from openpyxl import load_workbook #модуль для открытия существующего файла

def write_to_exel(row, num_flag, image_car, image_num, file_name):
  """ Функция записи данных в файл excel. если файла еще нет, то он создается.
  На вход row- список параметров
          num_flag - соответствует номер с картинки номеру из pdf или нет. False/True
          image_car, image_num - пути к картинкам машины и номера соответственно (строки)
          file_name - название файла, в который записывать данные
  На выход ничего не подается, файл формируется внутри функции
  """
  try:
    wb = openpyxl.load_workbook(file_name) #если файл уже создан, то загружаю его
    ws = wb.active # делаем лист активным

    ws.append(row) #записываю данные в активный лист

    ws[f'K{ws.max_row}'] = num_flag # записываю соответствует номер с картинки поставнолению или нет
    ws.row_dimensions[ws.max_row].height = 150 #устанавливаю высоту текущей строки исходя из размеров бОльшего изображения

    img_car = Image(image_car) # считываю как картинку изображение автомобиля
    img_num = Image(image_num) #считываю как картинку изображение номера

    img_num.height, img_num.width = 30, 100 # задаю размеры изображения номера
    img_car.height, img_car.width = 200, 300 # задаю размеры изображения машины

    ws.add_image(img_car, f'I{ws.max_row}')
    ws.add_image(img_num, f'J{ws.max_row}')

    wb.save(file_name) #сохраняю файл

  except: #если файла еще не существует
    #print('Создаю файл')
    wb = Workbook() #создаю книгу эксель
    ws = wb.active # делаем единственный лист активным

    ws.title = 'Drivers' #присваиваю имя листу

    ws.row_dimensions[1].height = 30 #задаю высоту строки заголовка

    cells = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'] #Перечень столбцов заголовка
    width = [20, 25, 10, 10, 60, 15, 10, 15, 45, 20, 20, 10] #список велчин ширины для столбцов

    #список наименований столбцов
    names = ['Номер файла', 'Постановление', 'Дата', 'Время', 'Адрес', 'Гос.номер', 'Сумма штрафа', 'Номер СТС', 'Фото ТС', 'Фото гос.номера', 'Фото гос.номера соотв.(True/False)', 'ИГР знак']

    #Формирую шапку таблицы
    for i, c in enumerate(cells): #прохожу циклом по всем номерам ячеек заголовка
      ws.column_dimensions[c].width = width[i] #задаю ширину столбца
      ws[f'{c}1'] = names[i] #присваиваю ячейке имя из списка имен заголовка
      ws[f'{c}1'].font = Font(bold = True) #делаю текст ячейки жирным
      ws[f'{c}1'].alignment = Alignment(wrap_text=True) #включаю перенос текста в ячейке

    ws.append(row) #записываю данные в активный лист
    ws[f'K{ws.max_row}'] = num_flag # записываю соответствует номер с картинки поставнолению или нет

    ws.row_dimensions[ws.max_row].height = 150 #устанавливаю высоту текущей строки исходя из размеров бОльшего изображения

    img_car = Image(image_car) # считываю как картинку изображение автомобиля
    img_num = Image(image_num) #считываю как картинку изображение номера

    img_num.height, img_num.width = 30, 100 # задаю размеры изображения номера
    img_car.height, img_car.width = 200, 300 # задаю размеры изображения машины

    ws.add_image(img_car, f'I{ws.max_row}')
    ws.add_image(img_num, f'J{ws.max_row}')

    ws.freeze_panes = ws['A2'] #закрепляю верхнюю строку

    wb.save(file_name) #сохраняю файл
  return

def pars_pdf(pdf_file):
  """Функция разбора pdf по шаблону.
  На вход название pdf файла.
  На выход список выбранных данных и строки-пути к файлу с картинкой
  """
  with open(pdf_file, 'rb') as file: # считывание информации из pdf при помощи библиотеки PyPDF2
    pdf = PdfReader(file)

    page = pdf.pages[0] #открываю 1-ю страницу
    text = page.extract_text() #извлекаю текст из 1-й страницы

    res_num = re.search(r'ПОСТАНОВЛЕНИЕ\s*\d*?\s', text).group(0) #вычленяю кусок текста с номером постановления
    res_num = re.search(r'\d+', res_num).group(0) #вычленяю номер постановления
    #print(f'res_num - {res_num}')

    date = re.search(r'\d{2}\.\d{2}\.\d{4}\s*в\s*\d{2}:\d{2}:\d{2}', text).group(0) #вычленяю кусок текста с датой и временем нарушения
    #print(date)
    vi_date = re.split(r'\s*в\s*', date)[0] #запоминаю дату нарушения
    #print(vi_date)
    vi_time = re.split(r'\s*в\s*', date)[1] #зампоминаю время нарушения
    #print(vi_time)

    adr = re.search(r'по адресу\s*[\w\W\s\+\-"/\.,]*?обл.', text).group(0) #вычленяю кусок текста с адресом нарушения
    adr = re.sub('по адресу',' ', adr).strip(' ') #выбираю именно адрес нарушения
    #print(adr)

    try: # если структура номера автомобиля
      reg_num = re.search(r'знак\s*\w\d{3}\w{2}\d{2,3}', text).group(0) #вычленяю кусок с гос номером авто типа Н114ТН716 (если это не прицеп, у них другая структура номера)
      reg_num =  re.sub('знак', ' ', reg_num).strip(' ') #выбираю сам номер

    except: # если структура номера прицепа
      reg_num = re.search(r'знак\s*\w{2}\d{4}\d{2}', text).group(0) #вычленяю кусок с гос номером если это номер прицепа (ВУ 1238 16)
      reg_num =  re.sub('знак', ' ', reg_num).strip(' ') #выбираю сам номер

    fine = re.search(r'в размере\s*[\d\s\.\,]*\s*руб.', text).group(0) #вычленяю кусок с размером штрафа
    fine = re.sub('в размере', ' ',fine)
    fine = re.sub('руб.', ' ', fine).strip(' ')
    #print(fine)

    page = pdf.pages[1] #разбираем 2-ю страницу
    text = page.extract_text() #извлекаю текст из 2-й страницы

    sts = re.search(r'СТС:\s*\d*', text).group(0) #выбираю кусок текста с номером СТС
    sts = re.sub(r'СТС:\s*', ' ', sts).strip(' ') #выбираю номер СТС
    #print(sts)

    #count = 0 #счетчик изображений
    for image in page.images:
        if f'{image.name}' == 'img3.jpg': #проверяю по имени изображения что это картинка номера машины
                                          #картинки номера машины всегда img3

          with open(f'{image.name}', 'wb') as fp:
            fp.write(image.data) #печатаю в файл чтобы потом в эксель передать путь к этому файлу

        if f'{image.name}' == 'img4.jpg': #проверяю по имени изображения что это картинка самой машины
                                          #картинки машины всегда img4

          with open(f'{image.name}', 'wb') as fp:
            fp.write(image.data) #печатаю в файл чтобы потом в эксель передать путь к этому файлу
    #img3 - номер
    #img4 - машина
  return [res_num, vi_date, vi_time, adr, reg_num, fine, sts] #собираю в кучку добытые из pdf данные

def generate_report():
  """Функция создания файла отчета.
  Заходит в папку, проверяет, что файл создан сегодня,
  запоминает имя файла, вызывает функцию разбора pdf и записи выбранных дайнных в файл.
  На вход ничего, на выход ничего
  """
  for pdf_file in glob.iglob(f'**/*.pdf', recursive = True): #перебираю файлы в папке
    filetime = datetime.fromtimestamp(os.path.getctime(pdf_file)) #получаем время создания файла
    if filetime.date() == datetime.now().date(): # если файл создан сегодня, то продолжаю
      pdf_num = os.path.basename(pdf_file) #получаю номер файла
      print(pdf_num) #для инфо печатаю номер текущего pdf файла

      row = pars_pdf(pdf_file) #вызываю функцию разбора pdf, результат присваиваю переменной row
      row.insert(0,pdf_num) #добавляю название файла в начало списка


      #записываю результат разбора текущего pdf файоа в эксель файл
      write_to_exel(row, num_flag = 'Да', image_car = 'img4.jpg', image_num = 'img3.jpg', file_name = r'drivers.xlsx')
  return

#Вызов основной функции генераии отчета
generate_report()

#если нужно отправить отчет на почту
def send_email(report_path, recipient_email):
    """Отправляю файл по пути report_path на почту recipient_email
    """
    sender_email = '*****' #почта откуда отправить
    sender_password = '****' #пароль для доступа к почте и оптравления писем

    msg = EmailMessage()
    msg['Subject'] = 'Daily report'
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg.set_content('Во вложении ежедневный отчет')

    with open(report_path, 'r') as file:
        report_data = file.read()
        report_name = os.path.basename(report_path)
        msg.add_attachment(report_data, maintype = 'application', subtype = 'octet-stream', filename = report_name)

    with smtplib.SMTP_SSL('smtp.mail.ru', 465) as smtp:
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)

def daily_report():
    """Функция генерации отчета и его отправки на почту
    """
    generate_report() #запуск функции генерации отчета
    send_email(r'drivers.xlsx', '****') #функция отправки отчета (файл для отправки и почта куда отправить)